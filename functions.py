import openpyxl
from aiogram.types import Message



wb = openpyxl.load_workbook('grades.xlsx')
sheet = wb.active


subjects = [cell.value for cell in sheet[1][1:]]

#возвращает словарь с оценками студента с заданным именем
def get_grades(name):
    grades = {}
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == name:
            for i in range(1, len(subjects)+1):
                grades[subjects[i-1]] = row[i].value
            break
    return grades

#возвращает список всех учеников.
def get_students():
    students = []
    for row in sheet.iter_rows(min_row=2):
        students.append(row[0].value)
    return students

async def welcome(message: Message):
    await message.reply("Привет! Я бот для просмотра оценок учеников. Введите /help для помощи")

async def help(message: Message):
    students = get_students()
    response = "Список учеников и соответствующие команды для просмотра их оценок:\n"
    for student in students:
        response += f"/{student}\n"
    await message.reply(response)

#Возвращает оценки студента по всем предметам, если такой студент есть в списке
async def get_student_grades(message: Message):
    name = message.text[1:]
    grades = get_grades(name)
    if not grades:
        await message.reply("Студент не найден")
    else:
        response = f"{name}:"
        for subject, grade in grades.items():
            response += f"\n{subject} - {grade}"
        await message.reply(response)
